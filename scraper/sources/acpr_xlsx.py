"""
ACPR/REGAFI XLSX parser (2026 export).

Parses the official "Liste des organismes d'assurance actifs" Excel file,
which has 3 sheets:
  1. Organismes_assurance_DD.MM.YYYY  - main French entities (~632)
  2. Passeports_entrants_DD.MM.YYYY   - foreign entities under LPS (skipped by default)
  3. Vehicules_de_groupe_DD.MM.YYYY   - SGA/SGAM/SGAPS/UMG group vehicles (~82)
"""
import logging
from pathlib import Path

from openpyxl import load_workbook

from sources.base import make_entity_dict

logger = logging.getLogger(__name__)


# Mapping FORME_JURIDIQUE + ACTIVITE -> normalized type_organisme
# Note: the source file has Latin-1 mojibake for accented chars (é -> � in some fields).
# We classify primarily by FORME_JURIDIQUE which is ASCII-safe.
def _classify(sous_categorie: str, forme: str, activite: str) -> tuple:
    """Returns (type_organisme, category)."""
    fj = (forme or "").upper().strip()
    act = (activite or "").lower().strip()

    # 1) Mutuelle: FORME = MUT or "MUT" prefix
    if fj == "MUT" or fj.startswith("MUT"):
        return "mutuelle", "mutuelle_sante" if act == "non-vie" else "mutuelle"

    # 2) Institution de prevoyance: FORME = IP
    if fj == "IP":
        return "institution_prevoyance", "institution_prevoyance"

    # 3) Reassurance
    if (act.startswith("r") and "assuranc" in act) or fj.startswith("R"):
        if (act.startswith("r") and "assuranc" in act):
            return "reassurance", "reassurance"

    # 4) Assurances: classify by ACTIVITE
    if act == "vie":
        return "assurance_vie", "assurance_vie"
    if act == "non-vie":
        return "assurance_non_vie", "assurance_non_vie"
    if act == "mixte":
        return "assurance_mixte", "assurance_mixte"
    if (act.startswith("r") and "assuranc" in act):
        return "reassurance", "reassurance"

    return "autre", "autre"


def _classify_vehicule(forme: str) -> tuple:
    """Vehicule de groupe (SGA, SGAM, SGAPS, UMG, ...)"""
    fj = (forme or "").upper().strip()
    return "groupe", f"vehicule_groupe_{fj.lower()}" if fj else "vehicule_groupe"


def _clean(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_acpr_xlsx(xlsx_path: str, include_lps: bool = False) -> list:
    """Parse the REGAFI XLSX file. Returns list of normalized entity dicts."""
    p = Path(xlsx_path)
    if not p.exists():
        logger.error(f"XLSX not found: {xlsx_path}")
        return []

    wb = load_workbook(str(p), read_only=True, data_only=True)
    entities = []
    seen_ids = set()

    # --- Sheet 1: Organismes_assurance_*
    main_sheets = [s for s in wb.sheetnames if s.lower().startswith("organismes")]
    for sn in main_sheets:
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        # Find column indexes
        idx = {h: i for i, h in enumerate(header) if h}

        def col(row, name, default=""):
            i = idx.get(name)
            return _clean(row[i]) if i is not None and i < len(row) else default

        for row in rows:
            if not row or not any(row):
                continue
            denom = col(row, "DENOMINATION_SOCIALE")
            siren = col(row, "SIREN")
            if not denom or not siren:
                continue
            sous_cat = col(row, "SOUS_CATEGORIE")
            forme = col(row, "FORME_JURIDIQUE")
            activite = col(row, "ACTIVITE")
            type_norm, category = _classify(sous_cat, forme, activite)

            ent = make_entity_dict(
                denomination=denom,
                siren=siren,
                matricule=col(row, "Id_REFASSU"),
                type_organisme=type_norm,
                category=category,
                address_street=col(row, "ADRESSE_SIEGE_SOCIAL"),
                postal_code=col(row, "CODE_POSTAL"),
                city=col(row, "VILLE"),
                source="acpr_xlsx_2026",
                source_url="https://acpr.banque-france.fr/registre-officiel-des-organismes-dassurance",
            )
            # Extra fields
            ent["lei"] = col(row, "LEI")
            ent["forme_juridique"] = forme
            ent["activite_acpr"] = activite
            ent["sous_categorie"] = sous_cat
            ent["branches_agrements"] = col(row, "BRANCHES_AGREMENTS")
            ent["sous_etat"] = col(row, "SOUS_ETAT")

            if ent["id"] in seen_ids:
                continue
            seen_ids.add(ent["id"])
            entities.append(ent)
        logger.info(f"  {sn}: {len(entities)} cumulative")

    # --- Sheet 3: Vehicules_de_groupe_* (sheet name has mojibake "V�hicules")
    veh_sheets = [s for s in wb.sheetnames
                  if "groupe" in s.lower() and "passeport" not in s.lower()
                  and "organisme" not in s.lower()]
    for sn in veh_sheets:
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        idx = {h: i for i, h in enumerate(header) if h}

        def col(row, name, default=""):
            i = idx.get(name)
            return _clean(row[i]) if i is not None and i < len(row) else default

        for row in rows:
            if not row or not any(row):
                continue
            denom = col(row, "DENOMINATION_SOCIALE")
            siren = col(row, "SIREN")
            if not denom or not siren:
                continue
            forme = col(row, "FORME_JURIDIQUE")
            type_norm, category = _classify_vehicule(forme)

            ent = make_entity_dict(
                denomination=denom,
                siren=siren,
                matricule=col(row, "Id_REFASSU"),
                type_organisme=type_norm,
                category=category,
                address_street=col(row, "ADRESSE_SIEGE_SOCIAL"),
                postal_code=col(row, "CODE_POSTAL"),
                city=col(row, "VILLE"),
                source="acpr_xlsx_2026",
                source_url="https://acpr.banque-france.fr/registre-officiel-des-organismes-dassurance",
            )
            ent["lei"] = col(row, "LEI")
            ent["forme_juridique"] = forme
            ent["sous_categorie"] = col(row, "SOUS_CATEGORIE")

            if ent["id"] in seen_ids:
                continue
            seen_ids.add(ent["id"])
            entities.append(ent)
        logger.info(f"  {sn}: {len(entities)} cumulative")

    # --- Sheet 2: LPS (foreign entities) - optional
    if include_lps:
        lps_sheets = [s for s in wb.sheetnames if s.lower().startswith("passeport")]
        for sn in lps_sheets:
            ws = wb[sn]
            rows = ws.iter_rows(values_only=True)
            header = next(rows)
            idx = {h: i for i, h in enumerate(header) if h}

            def col(row, name, default=""):
                i = idx.get(name)
                return _clean(row[i]) if i is not None and i < len(row) else default

            for row in rows:
                if not row or not any(row):
                    continue
                denom = col(row, "DENOMINATION_COMMERCIALE") or col(row, "DENOMINATION_SIEGE_SOCIAL")
                if not denom:
                    continue
                pays = col(row, "PAYS_D_ORIGINE")
                ent = make_entity_dict(
                    denomination=denom,
                    siren=col(row, "SIREN"),
                    type_organisme="autre",
                    category=f"lps_{pays.lower()}",
                    address_street=col(row, "ADRESSE_SIEGE_SOCIAL"),
                    source="acpr_xlsx_2026_lps",
                )
                ent["lei"] = col(row, "LEI_SIEGE_SOCIAL")
                ent["pays_origine"] = pays
                if ent["id"] in seen_ids:
                    continue
                seen_ids.add(ent["id"])
                entities.append(ent)
            logger.info(f"  {sn}: {len(entities)} cumulative (with LPS)")

    wb.close()
    logger.info(f"ACPR XLSX parsed: {len(entities)} entities")
    return entities


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO)
    sys.path.insert(0, "..")
    path = sys.argv[1] if len(sys.argv) > 1 else "data/acpr_2026.xlsx"
    ents = parse_acpr_xlsx(path)
    print(f"Total: {len(ents)}")
    types = {}
    for e in ents:
        types[e["type_organisme"]] = types.get(e["type_organisme"], 0) + 1
    for t, c in sorted(types.items(), key=lambda x: -x[1]):
        print(f"  {t}: {c}")
